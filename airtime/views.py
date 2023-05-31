from django.http import HttpResponseRedirect
from django.shortcuts import render
from random import randint
from .models import Airtime, Gfamily, DAirtime, AirtimeReceive
from .forms import AirtimeForm
import openpyxl
import json
import requests
from datetime import datetime


#define the http request that a user need to request 
"""def index(request):
    return render(request, 'airtime/index.html', {
        'airtimes' : Airtime.objects.all()
    })


def view_airtime(request, id):
    airtime = Airtime.objects.get(pk=id)
    return HttpResponseRedirect(reversed('index'))

    """

def index(request):
    return render(request, 'airtime/index.html', {
        'airtimesd' : AirtimeReceive.objects.all()
    })


def view_airtime(request, airtime_number):
    airtime = AirtimeReceive.objects.get(pk=airtime_number)
    return HttpResponseRedirect(reversed('index'))

def upload(request):
    if "GET" == request.method:    
        return render(request, 'airtime/uploadexcel.html', {})
    else:
        now = datetime.now()
        get_date_time = now.strftime("%d-%m-%Y %H:%M")

        at_airtime_url = 'https://api.africastalking.com/version1/airtime/send'
        dtree_request_body = {"username":"muchbeerapi"}
        randomvalue = randint(10, 100000)
        
        at_headers = {"Content-Type": "application/json", "Accept":"application/json" , 
                      "apiKey": "a2c23698253eafef3f02254cc9c414712ea0cad522ce1bd531ad126c353d959f", 
                      "Idempotency-Key": "dtree{}".format(randomvalue)}
        print(f"header printing2: {at_headers}")
        
        # try put validations here to check extension 

        #excel load data function here
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        
        worksheet = wb["Sheet1"]
        print(worksheet)

        g_families = []
        checkValues = False
        checkRetryValue = False
        #dtree_airtime_sdk = DAirtime().send()

        # getting value from each cell in row
        for rows in range(2, worksheet.max_row + 1):
            g_phoneNumber = worksheet.cell(rows, 1).value
            g_amount = worksheet.cell(rows, 2).value

            at_object = Gfamily(phoneNumber= "+{}".format(g_phoneNumber), amount="TZS {}".format(g_amount))
            at_object_to_dict = at_object.__dict__
            
            g_families.append(at_object_to_dict)
            
        dtree_request_body["recipients"] = g_families
    
        dtree_json_request = json.dumps(dtree_request_body, indent=4)
        print(f"The true request is : {dtree_json_request}")

        dtree_response = requests.post(at_airtime_url, data=dtree_json_request, headers=at_headers)
        
        dtree_response_dict = dtree_response.json()
        airtime_sent_error_message = dtree_response_dict.get("errorMessage")
        airtime_sent_response = dtree_response_dict.get("responses")
        print(f"error message is : {airtime_sent_error_message}")
        
        
        if(airtime_sent_error_message == "None"):
            """ dtree_random_super = f"batch{randint(500,200000)}"
            airtime_main = DAirtimeMain(dtree_date=get_date_time, total_amount= dtree_response_dict.get("totalAmount"), 
                                        special_value=dtree_random_super    )
            airtime_main.save()
            """
            checkValues = True

            for airtime in airtime_sent_response:
                airtime_received = AirtimeReceive(
                phoneNumber=airtime.get("phoneNumber"), amount= airtime.get("amount"), errorMessage=airtime.get("errorMessage"),status=airtime.get("status"))
                airtime_received.save()
        else:
            checkRetryValue = True

        return render(request, 'airtime/uploadexcel.html', { 
            'g_families':g_families,
            'dtree_response_dict': checkValues,
            'check_error': checkRetryValue})
    


def add(request):
    if request.method == 'POST':
        form = AirtimeForm(request.POST)
        if form.is_valid():
            new_airtime_number = form.cleaned_data['airtime_number']
            new_phone_number = form.cleaned_data['phone_number']
            new_amount = form.cleaned_data['amount']
            new_username = form.cleaned_data['username']

            new_airtime = Airtime(
                airtime_number = new_airtime_number,
                phone_number = new_phone_number,
                amount = new_amount,
                username = new_username
            )
                
            new_airtime.save()
            return render(request, 'airtime/add.html', {
                'form' : AirtimeForm(),
                'success': True
            })
    else:
        form = AirtimeForm()
    return render(request, 'airtime/add.html', {
        'form': AirtimeForm()
    })